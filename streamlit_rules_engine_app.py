import io
from datetime import date, datetime
from typing import Dict, List, Tuple

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


st.set_page_config(page_title="🦟 Malaria Excel Rule Engine", layout="wide")


# -----------------------------
# Helpers
# -----------------------------
def split_pipe(value) -> List[str]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return []
    text = str(value).strip()
    if not text:
        return []
    return [part.strip() for part in text.split("|") if part.strip()]


def is_blank(value) -> bool:
    return value is None or (isinstance(value, float) and pd.isna(value)) or str(value).strip() == ""


def row_to_excel_safe(row):
    out = []
    for value in row:
        if pd.isna(value):
            out.append(None)
        elif isinstance(value, pd.Timestamp):
            out.append(value.to_pydatetime())
        else:
            out.append(value)
    return out


def to_excel_safe(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out = out.where(pd.notna(out), None)
    for col in out.columns:
        if str(col).strip().casefold() == "screening_date":
            def _as_date(v):
                if v is None or pd.isna(v):
                    return None
                if isinstance(v, pd.Timestamp):
                    return v.date()
                if isinstance(v, datetime):
                    return v.date()
                return v

            out[col] = out[col].map(_as_date)
    return out


def put_comment_first(df: pd.DataFrame) -> pd.DataFrame:
    return df if "COMMENT" not in df.columns else df[["COMMENT"] + [c for c in df.columns if c != "COMMENT"]]


def _df_display_without_time(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col, dtype in out.dtypes.items():
        if pd.api.types.is_datetime64_any_dtype(dtype):
            out[col] = pd.to_datetime(out[col], errors="coerce").dt.strftime("%d-%b-%Y")
    return out


def _build_error_summary_counts(comment_series: pd.Series) -> pd.DataFrame:
    counts = {}
    for comment in comment_series.astype(str):
        parts = [p.strip() for p in comment.split(";") if p.strip()]
        for msg in parts:
            msg_lower = msg.lower()
            if msg_lower.endswith(" missing"):
                key = "Missing"
            elif msg_lower.endswith(" invalid option"):
                key = "Invalid option"
            elif msg_lower.endswith(" invalid number"):
                key = "Invalid number"
            elif msg_lower.endswith(" out of range"):
                key = "Out of range"
            elif msg_lower.endswith(" invalid format"):
                key = "Date invalid format"
            elif "inconsistent" in msg_lower:
                key = "Inconsistent"
            else:
                key = "Other"
            counts[key] = counts.get(key, 0) + 1

    if not counts:
        return pd.DataFrame(columns=["Error Type", "Count"])

    return (
        pd.DataFrame(list(counts.items()), columns=["Error Type", "Count"])
        .sort_values("Count", ascending=False)
        .reset_index(drop=True)
    )


def _build_error_detail_counts(comment_series: pd.Series) -> pd.DataFrame:
    counts = {}
    for comment in comment_series.astype(str):
        parts = [p.strip() for p in comment.split(";") if p.strip()]
        for msg in parts:
            counts[msg] = counts.get(msg, 0) + 1

    if not counts:
        return pd.DataFrame(columns=["Error Type Detail", "Count"])

    return (
        pd.DataFrame(list(counts.items()), columns=["Error Type Detail", "Count"])
        .sort_values("Count", ascending=False)
        .reset_index(drop=True)
    )


def _style_export_sheet(ws):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    widths = {}
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            cell_text = "" if value is None else str(value)
            base_max = 60 if idx == 1 else 24
            widths[idx] = min(max(widths.get(idx, 0), len(cell_text) + 2), base_max)
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def parse_date_value(value, formats: List[str]):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None, "blank"
    if isinstance(value, pd.Timestamp):
        return value.date(), None
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None

    text = str(value).strip()
    if not text:
        return None, "blank"

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date(), None
        except ValueError:
            pass
    return None, "invalid"


def compare_values(left, operator: str, right) -> bool:
    operator = operator.strip().upper()
    if operator in {"==", "EQ", "EQUALS"}:
        return str(left).strip() == str(right).strip()
    if operator in {"!=", "NE", "NOT_EQUALS"}:
        return str(left).strip() != str(right).strip()
    return False


def load_rules_from_workbook(rule_bytes: bytes):
    xls = pd.ExcelFile(io.BytesIO(rule_bytes))
    required_sheets = {"Rules", "Mappings"}
    missing_sheets = required_sheets - set(xls.sheet_names)
    if missing_sheets:
        raise ValueError(f"Rule workbook is missing required sheet(s): {', '.join(sorted(missing_sheets))}")

    rules = pd.read_excel(io.BytesIO(rule_bytes), sheet_name="Rules")
    mappings = pd.read_excel(io.BytesIO(rule_bytes), sheet_name="Mappings")

    needed_cols = {
        "rule_set", "priority", "active", "rule_type", "target_column", "source_column",
        "condition_column", "condition_operator", "condition_value", "expected_value",
        "allowed_values", "min_value", "max_value", "min_inclusive", "max_inclusive",
        "date_formats", "date_min", "date_max", "derived_column", "mapping_name",
        "comment_template"
    }
    missing_cols = needed_cols - set(rules.columns)
    if missing_cols:
        raise ValueError(f"Rules sheet missing column(s): {', '.join(sorted(missing_cols))}")

    mapping_needed = {"mapping_name", "source_value", "mapped_value"}
    missing_map_cols = mapping_needed - set(mappings.columns)
    if missing_map_cols:
        raise ValueError(f"Mappings sheet missing column(s): {', '.join(sorted(missing_map_cols))}")

    rules = rules.fillna("")
    mappings = mappings.fillna("")
    rule_sets = sorted({str(v).strip() for v in rules["rule_set"] if str(v).strip()})
    return rules, mappings, rule_sets


def build_mapping_lookup(mappings_df: pd.DataFrame) -> Dict[str, Dict[str, str]]:
    output: Dict[str, Dict[str, str]] = {}
    for _, row in mappings_df.iterrows():
        name = str(row["mapping_name"]).strip()
        if not name:
            continue
        output.setdefault(name, {})[str(row["source_value"]).strip()] = str(row["mapped_value"]).strip()
    return output


def apply_rule_set(df: pd.DataFrame, rules_df: pd.DataFrame, mappings_df: pd.DataFrame, rule_set: str) -> pd.DataFrame:
    out = df.copy()
    if "COMMENT" not in out.columns:
        out["COMMENT"] = ""

    selected_rules = (
        rules_df[
            (rules_df["rule_set"].astype(str).str.strip() == rule_set)
            & (rules_df["active"].astype(str).str.upper().isin(["Y", "YES", "TRUE", "1"]))
        ]
        .copy()
        .sort_values("priority")
    )
    mapping_lookup = build_mapping_lookup(mappings_df)
    parsed_dates: Dict[str, pd.Series] = {}

    for _, rule in selected_rules.iterrows():
        rule_type = str(rule["rule_type"]).strip().lower()
        target_col = str(rule["target_column"]).strip()
        source_col = str(rule["source_column"]).strip()
        derived_col = str(rule["derived_column"]).strip()
        comment_template = str(rule["comment_template"]).strip()

        if rule_type in {"missing", "choice", "numeric", "range", "date_format", "date_range", "consistency_if_equals"}:
            if target_col and target_col not in out.columns:
                out[target_col] = pd.NA
        if rule_type == "derive_map" and derived_col and derived_col not in out.columns:
            out[derived_col] = pd.NA

        if rule_type == "missing":
            mask = out[target_col].map(is_blank)
            if mask.any():
                out.loc[mask, "COMMENT"] = out.loc[mask, "COMMENT"].astype(str) + (comment_template + "; ")

        elif rule_type == "choice":
            allowed = split_pipe(rule["allowed_values"])
            mask = (~out[target_col].map(is_blank)) & (~out[target_col].astype(str).str.strip().isin(allowed))
            if mask.any():
                out.loc[mask, "COMMENT"] = out.loc[mask, "COMMENT"].astype(str) + (comment_template + "; ")

        elif rule_type == "numeric":
            parsed = pd.to_numeric(out[target_col], errors="coerce")
            mask = (~out[target_col].map(is_blank)) & parsed.isna()
            if mask.any():
                out.loc[mask, "COMMENT"] = out.loc[mask, "COMMENT"].astype(str) + (comment_template + "; ")

        elif rule_type == "range":
            parsed = pd.to_numeric(out[target_col], errors="coerce")
            valid_numeric = (~out[target_col].map(is_blank)) & parsed.notna()
            min_value = pd.to_numeric(pd.Series([rule["min_value"]]), errors="coerce").iloc[0]
            max_value = pd.to_numeric(pd.Series([rule["max_value"]]), errors="coerce").iloc[0]
            min_inclusive = str(rule["min_inclusive"]).strip().upper() in {"Y", "YES", "TRUE", "1"}
            max_inclusive = str(rule["max_inclusive"]).strip().upper() in {"Y", "YES", "TRUE", "1"}
            mask = pd.Series(False, index=out.index)
            if pd.notna(min_value):
                mask = mask | (parsed < min_value if min_inclusive else parsed <= min_value)
            if pd.notna(max_value):
                mask = mask | (parsed > max_value if max_inclusive else parsed >= max_value)
            mask = mask & valid_numeric
            if mask.any():
                out.loc[mask, "COMMENT"] = out.loc[mask, "COMMENT"].astype(str) + (comment_template + "; ")

        elif rule_type == "date_format":
            formats = split_pipe(rule["date_formats"])
            parsed_values = []
            invalid_mask = []
            for value in out[target_col]:
                parsed, status = parse_date_value(value, formats)
                parsed_values.append(parsed)
                invalid_mask.append(status == "invalid")
            parsed_dates[target_col] = pd.Series(parsed_values, index=out.index)
            if any(invalid_mask):
                mask = pd.Series(invalid_mask, index=out.index)
                out.loc[mask, "COMMENT"] = out.loc[mask, "COMMENT"].astype(str) + (comment_template + "; ")
            # replace valid entries with true date objects for export
            out[target_col] = parsed_dates[target_col].where(parsed_dates[target_col].notna(), out[target_col])

        elif rule_type == "date_range":
            if target_col not in parsed_dates:
                formats = split_pipe(rule["date_formats"])
                parsed_dates[target_col] = out[target_col].map(lambda v: parse_date_value(v, formats)[0])
            min_date = None if str(rule["date_min"]).strip() == "" else date.fromisoformat(str(rule["date_min"]).strip())
            max_text = str(rule["date_max"]).strip().upper()
            max_date = date.today() if max_text == "TODAY" else (None if not max_text else date.fromisoformat(str(rule["date_max"]).strip()))
            mask = parsed_dates[target_col].notna()
            if min_date is not None:
                mask = mask & (parsed_dates[target_col] < min_date)
            else:
                mask = pd.Series(False, index=out.index)
            if max_date is not None:
                mask = mask | ((parsed_dates[target_col].notna()) & (parsed_dates[target_col] > max_date))
            if mask.any():
                out.loc[mask, "COMMENT"] = out.loc[mask, "COMMENT"].astype(str) + (comment_template + "; ")

        elif rule_type == "consistency_if_equals":
            cond_col = str(rule["condition_column"]).strip()
            cond_op = str(rule["condition_operator"]).strip() or "=="
            cond_value = str(rule["condition_value"]).strip()
            expected_value = str(rule["expected_value"]).strip()
            applicable = out[cond_col].map(lambda x: compare_values(x, cond_op, cond_value))
            inconsistent = applicable & (out[target_col].astype(str).str.strip() != expected_value)
            if inconsistent.any():
                out.loc[inconsistent, "COMMENT"] = out.loc[inconsistent, "COMMENT"].astype(str) + (comment_template + "; ")

        elif rule_type == "derive_map":
            mapping_name = str(rule["mapping_name"]).strip()
            mapper = mapping_lookup.get(mapping_name, {})
            out[derived_col] = out[source_col].map(lambda v: mapper.get(str(v).strip(), pd.NA) if not is_blank(v) else pd.NA)

    out["COMMENT"] = out["COMMENT"].astype(str).str.strip().str.rstrip(";").str.strip()
    return out


def build_export_workbook(processed: Dict[str, pd.DataFrame], original_bytes: bytes, sheet_summaries: Dict[str, Dict[str, pd.DataFrame]]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    orig_wb = load_workbook(io.BytesIO(original_bytes))

    if "Error Summary" in wb.sheetnames:
        ws_sum = wb["Error Summary"]
    else:
        ws_sum = wb.create_sheet("Error Summary", 0)

    rows_type = []
    detail_counts = {}
    for sheet_name, bundle in sheet_summaries.items():
        df_type = bundle["type_summary"]
        for _, row in df_type.iterrows():
            rows_type.append([sheet_name, row["Error Type"], int(row["Count"])])
        df_detail = bundle["detail_summary"]
        for _, row in df_detail.iterrows():
            detail_counts[row["Error Type Detail"]] = detail_counts.get(row["Error Type Detail"], 0) + int(row["Count"])

    df_type_summary = pd.DataFrame(rows_type, columns=["Sheet", "Error Type", "Count"]) if rows_type else pd.DataFrame(columns=["Sheet", "Error Type", "Count"])
    df_detail_summary = (
        pd.DataFrame(list(detail_counts.items()), columns=["Error Type Detail", "Count"]) 
        .sort_values("Count", ascending=False)
        .reset_index(drop=True)
        if detail_counts else pd.DataFrame(columns=["Error Type Detail", "Count"])
    )

    def write_df(ws, df, title, start_row):
        ws.cell(row=start_row, column=1, value=title)
        start_row += 1
        for j, col in enumerate(df.columns, start=1):
            ws.cell(row=start_row, column=j, value=col)
        start_row += 1
        for row in df.itertuples(index=False):
            for j, value in enumerate(row, start=1):
                ws.cell(row=start_row, column=j, value=value)
            start_row += 1
        return start_row

    r = 1
    r = write_df(ws_sum, df_type_summary, "Error Type Summary", r) + 2
    write_df(ws_sum, df_detail_summary, "Error Type Detail", r)
    _style_export_sheet(ws_sum)

    for sheet_name, df in processed.items():
        title = f"Processed - {sheet_name}"
        ws = wb.create_sheet(title)
        safe_df = to_excel_safe(df)
        for row in dataframe_to_rows(safe_df, index=False, header=True):
            ws.append(row_to_excel_safe(row))
        _style_export_sheet(ws)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def main():
    st.title("🦟 Malaria Excel Rule Engine")
    st.caption("Sample version: rules are stored in Excel, and Python/Streamlit reads and executes them dynamically.")

    st.header("1) Upload rule workbook")
    rule_file = st.file_uploader("Upload rules workbook", type=["xlsx"], key="rule_workbook")

    rules_df = mappings_df = None
    rule_sets: List[str] = []

    if rule_file is not None:
        try:
            rule_bytes = rule_file.getvalue()
            rules_df, mappings_df, rule_sets = load_rules_from_workbook(rule_bytes)
            st.success(f"Loaded rule workbook with {len(rule_sets)} rule set(s).")
            with st.expander("Preview rule rows"):
                st.dataframe(rules_df.head(20), use_container_width=True)
        except Exception as exc:
            st.error("Failed to read the rule workbook.")
            st.exception(exc)
            st.stop()
    else:
        st.info("Upload the sample rules workbook first.")
        st.stop()

    rule_set = st.selectbox("Choose rule set", rule_sets)

    st.header("2) Upload malaria data workbook")
    data_file = st.file_uploader("Upload malaria Excel file", type=["xlsx", "xls"], key="data_workbook")

    if data_file is None:
        st.warning("Upload a data workbook to continue.")
        st.stop()
    sheet_names = []
    try:
        original_bytes = data_file.getvalue()
        xls = pd.ExcelFile(io.BytesIO(original_bytes))
        sheet_names = xls.sheet_names
    except Exception as exc:
        st.error("Could not read the uploaded data workbook.")
        st.exception(exc)
        st.stop()

    st.header("3) Choose sheet(s)")
    if len(sheet_names) == 1:
        selected = sheet_names[:]
        st.info(f"Only one sheet found: {sheet_names[0]}")
    else:
        multi = st.checkbox("Process multiple sheets", value=False)
        selected = st.multiselect("Select sheet(s)", sheet_names, default=sheet_names[:1]) if multi else [st.selectbox("Select sheet", sheet_names)]

    run_btn = st.button("Run validation")

    if run_btn:
        processed = {}
        summaries = {}
        errors: List[Tuple[str, Exception]] = []

        for sheet_name in selected:
            try:
                raw_df = xls.parse(sheet_name=sheet_name)
                out_df = apply_rule_set(raw_df, rules_df, mappings_df, rule_set)
                processed[sheet_name] = put_comment_first(out_df)
                comments = out_df.get("COMMENT", pd.Series("", index=out_df.index))
                summaries[sheet_name] = {
                    "type_summary": _build_error_summary_counts(comments),
                    "detail_summary": _build_error_detail_counts(comments),
                }
            except Exception as exc:
                errors.append((sheet_name, exc))

        if processed:
            st.subheader("Preview")
            tabs = st.tabs(list(processed.keys()))
            for tab, sheet_name in zip(tabs, processed.keys()):
                with tab:
                    st.dataframe(_df_display_without_time(processed[sheet_name].head(10)), use_container_width=True)

            st.subheader("Error summary")
            for sheet_name, bundle in summaries.items():
                st.markdown(f"**{sheet_name}**")
                st.dataframe(bundle["type_summary"], use_container_width=True)
                st.dataframe(bundle["detail_summary"], use_container_width=True)

            output_bytes = build_export_workbook(processed, original_bytes, summaries)
            st.download_button(
                label="📥 Download processed workbook",
                data=output_bytes,
                file_name="malaria_processed_with_summary.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        if errors:
            st.warning("Some sheets failed to process.")
            for sheet_name, exc in errors:
                with st.expander(f"Details: {sheet_name}"):
                    st.exception(exc)


if __name__ == "__main__":
    main()
